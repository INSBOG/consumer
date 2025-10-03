import json
from io import BytesIO
from uuid import uuid4

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from validators.general_validator import GeneralValidator
from validators.pat_type_validator import PatTypeValidator
from validators.spec_validator import SpecValidator
from validators.ward_validator import WardValidator
from validators.antibiotic_validator import (
    AntComparisonValidator,
    Ant,
    AntEnum,
    Hoja,
)
from validators.antibiotic_2_validator import Antibiotic2Validator
from openpyxl.cell import Cell
from constants import not_empty_columns


# Definir el estilo para celdas con errores
error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")


class ValidatorService:

    socket_service = None


    def __update_progress(self, progress: float, report: dict, message: str):
        if not self.socket_service:
            return
        self.socket_service.emit("progress", json.dumps({
            "progress": progress,
            "report": str(report["_id"]),
            "message": message
        }))

    def validate(self, data: BytesIO, report: dict):
        df = pd.read_excel(data, keep_default_na=False, na_values=[""])

        df.columns = df.columns.str.replace(r",(?:[A-Z](?:,[0-9]+)?)", "", regex=True)

        with open("data.json", "w") as f:
            f.write(df.head().to_json(orient="records"))

        df["VALIDATION_ERRORS"] = ""
        df["ANTIBIOTIC_ERRORS"] = ""
        df["MISSING_COLUMNS"] = ""

        missing_cols = [col for col in not_empty_columns if col not in df.columns.to_list()]
        df.at[0, "MISSING_COLUMNS"] = ";".join([f"Falta columna: {col}" for col in missing_cols])

        errors = {}
        ant_errors = {}

        items = [
            Ant("A:H", AntEnum.ETP_NM, Hoja.HOJA2),
            Ant("J:Q", AntEnum.IPM_NM, Hoja.HOJA2),
            Ant("S:Z", AntEnum.MEM_NM, Hoja.HOJA2),
            Ant("AB:AF", AntEnum.FEP_NM, Hoja.HOJA2),
            Ant("A:E", AntEnum.CZA_NM, Hoja.HOJA3),
            Ant("G:L", AntEnum.CZA_NM, Hoja.HOJA3),
        ]
        # Crear una lista de validadores
        validators = [
            GeneralValidator(df, errors),
            WardValidator(df, report, errors),
            PatTypeValidator(df, errors),
            SpecValidator(df, errors),
            Antibiotic2Validator(df, ant_errors),
            *[AntComparisonValidator(
                data=df,
                usecols=a.rng,
                ant=a.name,
                sheet_name=a.sheet,
                errors=ant_errors
            ) for a in items],
        ]

        for idx, validator in enumerate(validators, start=1):
            print(f"Validando {validator.__class__.__name__}...")
            self.__update_progress(idx / len(validators), report, f"Validando {validator.__class__.__name__}")
            validator.validate()

        wb = Workbook()

        ws = wb.active

        # Crear una copia del DataFrame original
        columns = df.columns.to_list()
        ws.append(columns)

        for index, row in df.iterrows():
            row_errors = errors.get(index, {})
            ant_row_errors = ant_errors.get(index, {})

            # Asignar errores consolidados a "VALIDATION_ERRORS"
            row["VALIDATION_ERRORS"] = "; ".join(row_errors.values())
            row["ANTIBIOTIC_ERRORS"] = "; ".join(set(ant_row_errors.values()))

            # Agregar fila completa al archivo Excel
            ws.append(row.to_list())

            # Procesar celdas con errores
            self.color_cell_with_error(df, ws, row_errors, index)
            self.color_cell_with_error(df, ws, ant_row_errors, index)

        # Colorizar columnas con errores
        self.color_column_with_errors({
            "ws": ws, 
            "columns": df.columns.to_list(), 
            "errors": errors, 
            "ant_errors": ant_errors
        })

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        wb.close()

        filename = f"{uuid4()}.xlsx"
        self.minio.bucket = "processed"
        self.minio.upload_file(filename, file_stream)
        self.__update_progress(5/5, report, "Reporte generado")

        return filename

    def color_column_with_errors(self, data: dict):
        ws = data.get("ws")
        columns = data.get("columns", [])
        errors = data.get("errors", {})
        ant_errors = data.get("ant_errors", {})

        # Fusionar errores antiguos con actuales
        for idx, error in ant_errors.items():
            errors.setdefault(idx, {}).update(error)

        # Crear un mapa de columnas a su índice (más rápido y seguro)
        column_indices = {col: i + 1 for i, col in enumerate(columns)}

        columnas_con_error = set()
        for idx, error in errors.items():
            for col in error:
                col_idx = column_indices.get(col)
                if col_idx:
                    columnas_con_error.add(col_idx)

        # Pintar encabezados de columnas con error
        for col_idx in columnas_con_error:
            header_cell = ws.cell(row=1, column=col_idx)
            header_cell.fill = error_fill


    def color_cell_with_error(self, df: pd.DataFrame, ws, errors: dict, index: int):
        for col, error in  errors.items():
            if col not in df.columns:
                continue

            col_idx = df.columns.get_loc(col) + 1
            cell = ws.cell(row=index+2, column=col_idx)
            cell.fill = error_fill
            cell.comment = Comment(error, "Validación")